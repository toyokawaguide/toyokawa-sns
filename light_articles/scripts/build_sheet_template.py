"""build_sheet_template.py — ライト記事キュー xlsx 最終版（v3）

社長要望全反映：
- 公開日列なし（自動判定）
- 元記事タイトル列追加（F列）
- 目印列追加（H列）
- サブ上書き列追加（L列）
- 14列構成
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

OUTPUT = Path(__file__).resolve().parent / "_sample" / "ライト記事キュー_最終版.xlsx"
OUTPUT.parent.mkdir(parents=True, exist_ok=True)

wb = Workbook()
ws = wb.active
ws.title = "キュー"

# 13列構成（D列パターン廃止・元記事タイトル有無で自動判定）
columns = [
    ("A", "ID",            12, "LR001 形式・連番"),
    ("B", "状態",           12, "draft / 予約済 / 公開済 / スキップ"),
    ("C", "場所",           28, "アイキャッチ左カード上段・タイトルにも使用"),
    ("D", "元記事URL",      36, "追跡元の過去記事URL（任意・本文に内部リンク挿入）"),
    ("E", "元記事タイトル", 50, "あり→続報モード／空→お知らせモード。長文は2段＋…で自動省略"),
    ("F", "住所",           28, "アイキャッチ右カード上段"),
    ("G", "目印",           24, "アイキャッチ右カード下段（例：国道151号線沿い）"),
    ("H", "地図URL",        30, "Google Maps URL（任意・本文に挿入）"),
    ("I", "タイトル上書き", 30, "WP記事タイトル上書き（空ならテンプレ自動生成）"),
    ("J", "本文上書き",     40, "WP本文上書き（空ならテンプレ自動生成）"),
    ("K", "サブ",           22, "アイキャッチ左カード下段「▼ その後」（必須・例：事務所として活用 / 閉店 / 移転）"),
    ("L", "メモ",           20, "社長メモ・自動投稿には使われない"),
    ("M", "担当",           10, "アヤ"),
]

# ヘッダー（占いと同じ紺）
header_fill = PatternFill("solid", fgColor="1A3A8A")
header_font = Font(bold=True, color="FFFFFF", size=11)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
border = Border(
    left=Side(style="thin", color="888888"),
    right=Side(style="thin", color="888888"),
    top=Side(style="thin", color="888888"),
    bottom=Side(style="thin", color="888888"),
)

for col_idx, (col_letter, name, width, comment) in enumerate(columns, start=1):
    cell = ws.cell(row=1, column=col_idx, value=name)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align
    cell.border = border
    cell.comment = Comment(comment, "アヤ")
    ws.column_dimensions[col_letter].width = width

# 例データ3行（実際の追跡候補をベースに）
examples = [
    {
        "A": "LR001",
        "B": "draft",
        "C": "とんやんラーメン跡地",
        "D": "https://toyokawa-rentallife.com/2024/12/...",
        "E": "【テナント・土地】だいぶ以前\"とんやんラーメン\"だったところがテナント募集してる。国道１５１号線沿い。ただし飲食店不可。",
        "F": "豊川市諏訪3丁目",
        "G": "国道151号線沿い",
        "H": "https://maps.app.goo.gl/...",
        "I": "",
        "J": "",
        "K": "事務所として活用",
        "L": "事務所判明",
        "M": "アヤ",
    },
    {
        "A": "LR002",
        "B": "draft",
        "C": "御油町イチトサンブンノイチ跡地",
        "D": "https://toyokawa-rentallife.com/2025/04/...",
        "E": "【閉店】御油町の「イチトサンブンノイチ」、9月オープンから半年で閉店していたみたい",
        "F": "豊川市御油町X-Y",
        "G": "御油の松並木近く",
        "H": "",
        "I": "",
        "J": "",
        "K": "新築事務所建設中",
        "L": "次は事務所予定",
        "M": "アヤ",
    },
    {
        "A": "LR003",
        "B": "draft",
        "C": "コトリコ栄町店",
        "D": "",  # 元記事URL なし → お知らせモード
        "E": "",  # 元記事タイトル なし → お知らせモード
        "F": "豊川市栄町X-Y",
        "G": "豊川駅徒歩3分",
        "H": "",
        "I": "",
        "J": "",
        "K": "閉店",  # K列「サブ」必須
        "L": "5/30 閉店",
        "M": "アヤ",
    },
]

data_align = Alignment(vertical="top", wrap_text=True)
for row_idx, ex in enumerate(examples, start=2):
    for col_idx, (col_letter, name, _, _) in enumerate(columns, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=ex.get(col_letter, ""))
        cell.alignment = data_align
        cell.border = border

ws.row_dimensions[1].height = 32
for r in range(2, 5):
    ws.row_dimensions[r].height = 90

# ===== 使い方シート =====
ws2 = wb.create_sheet("使い方")

notes = [
    ["■ ライト記事キュー 最終版 使い方"],
    [""],
    ["1. アヤが追跡記事ネタを見つけたら、この Sheets に行追加"],
    ["   - ID は LRxxx 連番（既存最大値+1）"],
    ["   - 状態は draft で開始"],
    ["   - 場所・パターン・元記事URL・元記事タイトル・住所・目印 を埋める"],
    ["   - L列「サブ上書き」は空でもOK（パターン既定が使われる）"],
    [""],
    ["2. 写真フォルダ準備（任意）"],
    ["   C:\\Users\\Yoshida\\Desktop\\豊川ガイド\\ライト記事\\{ID}_{場所}\\1.jpg, 2.jpg ..."],
    ["   例: LR001_とんやんラーメン跡地\\1.jpg"],
    ["   写真ゼロでもOK（テキストのみで公開）"],
    [""],
    ["■ 自動投稿の仕組み（毎日 18:55 JST cron）"],
    [""],
    ["Step 1: WP で「翌日19:00 公開予定の手動投稿」をチェック"],
    ["   - あり → スキップ（その日の自動投稿は休む）"],
    ["   - なし → Step 2 へ"],
    [""],
    ["Step 2: Sheets の draft 行を取得（行番号の小さい順＝古い順）"],
    ["   - draft が1つもなければ → 「ストック切れ」Gmail 通知"],
    [""],
    ["Step 3: 一番上の draft を翌日19:00 公開予約で WP に投稿"],
    ["   - アイキャッチ自動生成（場所・住所・目印・元記事タイトル使用）"],
    ["   - フォルダの番号付き写真を WPメディアアップロード→本文に挿入"],
    ["   - パターン別テンプレで本文生成（J列/K列の上書きがあれば優先）"],
    [""],
    ["Step 4: 状態を「予約済」に更新"],
    [""],
    ["Step 5: SNS自動投稿"],
    ["   - Threads / Instagram Feed / Instagram Reels"],
    [""],
    ["Step 6: Gmail通知（X予約用テキスト付き）"],
    ["   - スマホから X 公式アプリでコピペ予約"],
    [""],
    ["■ 手動投稿（社長が直接書く場合）"],
    ["この Sheets には入れない！"],
    ["WP管理画面で直接「お知らせ」カテゴリで19時予約 → 自動投稿が検知してその日はスキップ"],
    [""],
    ["■ モード判定（パターン列なし・E列で自動判定）"],
    [""],
    ["E列「元記事タイトル」", "→", "モード", "ラベル帯", "キャッチ", "カードラベル"],
    ["値あり", "→", "続報モード", "【続報】", "あの記事の答え合わせ", "【続報情報】"],
    ["空欄", "→", "お知らせモード", "【お知らせ】", "管理人のひとり言", "【お知らせ】"],
    [""],
    ["■ K列「サブ」記入例（必須）"],
    ["・テナント募集後→「事務所として活用」「賃貸住宅入居」"],
    ["・解体→建設中→「新築事務所建設中」「マンション工事中」"],
    ["・閉店→「閉店」"],
    ["・移転→「移転」"],
    ["・リニューアル→「リニューアル」"],
    [""],
    ["■ アイキャッチに表示される項目"],
    ["・C列 場所 → 【続報情報】カード上段"],
    ["・L列 サブ → 【続報情報】カード下段（▼その後）"],
    ["・G列 住所 → 【場所】カード上段"],
    ["・H列 目印 → 【場所】カード下段（▼目印）"],
    ["・F列 元記事タイトル → 「▼ 過去記事」枠（長文は2段＋…で自動省略）"],
    [""],
    ["■ 公開時間"],
    ["全て 19:00 JST 固定"],
    [""],
    ["■ X (Twitter) 運用"],
    ["毎日18:55 自動投稿時、Gmail に「X予約用テキスト」が届く"],
    ["スマホでコピー → X公式アプリの予約投稿で 翌日19:00 セット"],
    ["お風呂・通勤中などスキマ時間でOK"],
]

for r, row_data in enumerate(notes, start=1):
    for c, val in enumerate(row_data, start=1):
        cell = ws2.cell(row=r, column=c, value=val)
        if r == 1 or (val and isinstance(val, str) and (val.startswith("■") or val.startswith("Step "))):
            cell.font = Font(bold=True, color="1A3A8A", size=12)

ws2.column_dimensions["A"].width = 26
ws2.column_dimensions["B"].width = 40
ws2.column_dimensions["C"].width = 25

wb.save(OUTPUT)
print(f"✅ xlsx 最終版生成: {OUTPUT}")
print(f"   サイズ: {OUTPUT.stat().st_size / 1024:.1f} KB")
print(f"   列数: 14（A-N）")

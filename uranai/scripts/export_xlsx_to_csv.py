"""xlsx の各シートを CSV にエクスポート（Google Sheets 移行時の社長コピペ用）

実行：python export_xlsx_to_csv.py
出力先：uranai/data/csv_for_sheets/{ファイル名}_{シート名}.csv
"""
from __future__ import annotations
import sys, csv
from pathlib import Path
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
OUT = DATA / "csv_for_sheets"
OUT.mkdir(parents=True, exist_ok=True)

XLSX_FILES = [
    DATA / "ラッキースポット管理.xlsx",
    DATA / "生まれ年グループ.xlsx",
    DATA / "町名グループ.xlsx",
]

for xlsx in XLSX_FILES:
    if not xlsx.exists():
        print(f"[skip] {xlsx.name} not found")
        continue
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        # ファイル名は xlsxファイル名_シート名.csv
        out_file = OUT / f"{xlsx.stem}_{sheet_name}.csv"
        with open(out_file, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            for row in rows:
                # None を空文字に
                writer.writerow(["" if v is None else v for v in row])
        print(f"[OK] {out_file.relative_to(ROOT)} ({len(rows)} rows)")

print(f"\n出力先: {OUT}")
print("使い方: 各 CSV を開いて全選択→コピー→Google Sheet の該当タブに貼り付け")

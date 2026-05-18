"""Sheets→xlsx 自動同期（フォールバック用xlsxを常に最新Sheetsの写しに保つ）

【目的】
社長は Google Sheets だけ更新する。リポの xlsx は本来フォールバック用だが
社長更新と同期されないと「Sheets失敗時に古いデータで配信」する穴になる
（2026-05-18 障害の教訓）。本スクリプトを本配信(06:00)の前(05:45)に
cron 実行し、Sheets→xlsx を毎日同期＋commit することで
「Sheets失敗時のxlsxフォールバック＝最新Sheetsの写し」を保証する。

【フェイルセーフ（最重要）】
cron が読む5タブのうち1つでも取得失敗したら、xlsx を一切変更せず
非ゼロ終了する。部分更新で不整合を作るより、前回同期した完全な
状態を保持する方が安全（本配信は別workflowなので止まらない）。

【対象】cron が実際に読む5タブのみ（参考タブは触らない＝既存保持）
"""
from __future__ import annotations
import sys
import os
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"

# (Sheetsタブ名, xlsxファイル名, xlsxシート名)
# ※load_groups は xlsx シート名「配信スケジュール」を読むため、
#   Sheets「生まれ年配信スケジュール」→ xlsx「配信スケジュール」へ名前変換
SYNC_MAP = [
    ("ラッキースポット入力", "ラッキースポット管理.xlsx", "ラッキースポット入力"),
    ("マスタ", "ラッキースポット管理.xlsx", "マスタ"),
    ("配信ログ", "ラッキースポット管理.xlsx", "配信ログ"),
    ("生まれ年配信スケジュール", "生まれ年グループ.xlsx", "配信スケジュール"),
    ("町名配信スケジュール", "町名グループ.xlsx", "配信スケジュール"),
]


def main() -> None:
    import load_sheets

    if not load_sheets.is_enabled():
        print("[skip] URANAI_SHEETS_URL 未設定 → 同期しない（xlsx変更なし）")
        sys.exit(0)

    # === 1) 全タブ取得（1つでも失敗したら中止・xlsx触らない） ===
    fetched: dict[tuple[str, str], list] = {}
    for sheet_tab, xlsx_name, xlsx_sheet in SYNC_MAP:
        rows = load_sheets.fetch_sheet_normalized(sheet_tab)
        if rows is None:
            print(f"[FAIL] '{sheet_tab}' 取得失敗 → 同期中止"
                  f"（xlsx変更なし・前回の完全な状態を保持）")
            sys.exit(1)
        if not rows:
            print(f"[FAIL] '{sheet_tab}' が空 → 同期中止"
                  f"（空で上書きすると配信不能になるため保護）")
            sys.exit(1)
        fetched[(xlsx_name, xlsx_sheet)] = rows
        print(f"[OK] '{sheet_tab}' {len(rows)}行 取得")

    # === 2) 全タブ成功 → xlsxファイル別に該当シートだけ差し替え ===
    by_file: dict[str, list[tuple[str, list]]] = {}
    for (xlsx_name, xlsx_sheet), rows in fetched.items():
        by_file.setdefault(xlsx_name, []).append((xlsx_sheet, rows))

    for xlsx_name, sheets in by_file.items():
        path = DATA / xlsx_name
        if path.exists():
            wb = openpyxl.load_workbook(path)
        else:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

        for xlsx_sheet, rows in sheets:
            # 既存シートは位置を保ったまま中身だけ入れ替え
            if xlsx_sheet in wb.sheetnames:
                idx = wb.sheetnames.index(xlsx_sheet)
                del wb[xlsx_sheet]
                ws = wb.create_sheet(xlsx_sheet, idx)
            else:
                ws = wb.create_sheet(xlsx_sheet)
            for r in rows:
                ws.append(list(r))
            print(f"  -> {xlsx_name}[{xlsx_sheet}] {len(rows)}行 書込")

        # 参考タブ（使い方・候補リスト等）は触らず保持
        wb.save(path)
        print(f"[SAVE] {path.name}")

    print("=== Sheets→xlsx 同期完了 ===")


if __name__ == "__main__":
    main()

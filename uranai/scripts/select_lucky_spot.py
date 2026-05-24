"""
ラッキースポット選定（Phase 1 Step 3）
========================================

【ロジック】
1. 「ラッキースポット入力」シートで target_date を検索 → あればそれを採用（社長指名）
2. なければ「マスタ」シートから active なスポットをランダム選定
   - 直近7日に使用したスポットは除外（同週NG）
   - tier の偏り防止：直近7日の tier 集計で 50:30:20 から極端に離れた tier を弱く除外

【入出力】
- 入力: data/ラッキースポット管理.xlsx
  - 「ラッキースポット入力」シート（社長手動入力・優先）
  - 「マスタ」シート（フォールバック候補プール）
  - 「配信ログ」シート（直近7日の使用履歴）
- 出力: 選定結果 dict
  - {"date", "weekday", "name", "tier", "area", "memo", "source"}

【使い方】
  cd 占い/scripts

  # 単発選定
  python select_lucky_spot.py --date 2026-05-11

  # dry-run シミュレーション（5/4〜5/17 の14日分）
  python select_lucky_spot.py --simulate --start 2026-05-04 --end 2026-05-17

  # ライブラリとして使用
  from select_lucky_spot import select_lucky_spot
  result = select_lucky_spot(date(2026, 5, 11))
"""
from __future__ import annotations
import sys
import random
import argparse
from pathlib import Path
from datetime import date, datetime, timedelta
from collections import Counter
from dataclasses import dataclass, asdict

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
LUCKY_XLSX = ROOT / "data" / "ラッキースポット管理.xlsx"
INPUT_SHEET = "ラッキースポット入力"
MASTER_SHEET = "マスタ"
LOG_SHEET = "配信ログ"

WEEKDAY_JP = ["月", "火", "水", "木", "金", "土", "日"]


@dataclass
class Spot:
    name: str
    is_chain: bool = False     # チェーン店フラグ（True=「お近くの」付与）
    source: str = ""           # "input" / "master" / "fallback"
    # 互換用（既存コードが参照している場合に空文字を返す）
    tier: str = ""             # 内部分類用（is_chain で実質置換・将来削除可）
    area: str = ""
    memo: str = ""


def to_date(d) -> date:
    """文字列・datetime・date を date に正規化"""
    if isinstance(d, date) and not isinstance(d, datetime):
        return d
    if isinstance(d, datetime):
        return d.date()
    if isinstance(d, str):
        return datetime.strptime(d, "%Y-%m-%d").date()
    raise ValueError(f"unsupported date type: {type(d)}")


def load_workbook_ro():
    return openpyxl.load_workbook(LUCKY_XLSX, data_only=True)


def _get_rows(wb_or_none, sheet_name: str) -> list[tuple] | None:
    """Google Sheets からシートを取得（2026-05-25 案A：xlsx フォールバック完全削除）

    GHA 本番運用は Sheets 必須。ローカル CLI で wb を渡された時のみ xlsx を許可。
    """
    # Sheets 優先（GHA 本番ルート）
    try:
        import load_sheets
        if load_sheets.is_enabled():
            rows = load_sheets.fetch_sheet_normalized(sheet_name)
            if rows is not None:
                return rows
            # Sheets 有効だが取得失敗 → エラー（xlsx には落ちない）
            print(f"[error] Sheets から '{sheet_name}' 取得失敗。フォールバックなし")
            return None
    except ImportError:
        pass
    # Sheets 無効（ローカル CLI など）：wb が明示的に渡された時のみ xlsx 使用
    if wb_or_none is None:
        return None
    ws = wb_or_none[sheet_name]
    return list(ws.iter_rows(values_only=True))


def lookup_input_sheet(wb, target_date: date) -> Spot | None:
    """入力シートで target_date のスポットを検索

    シート構造（3列・2026-05-07 シンプル化版）:
        A: 配信日
        B: スポット名
        C: チェーン店?（空欄でない=チェーン店扱い）
    """
    rows = _get_rows(wb, INPUT_SHEET)
    if rows is None:
        return None
    # データ部は行6（index 5）以降
    for r in rows[5:]:
        if not r or r[0] is None:
            continue
        try:
            row_date = to_date(r[0])
        except (ValueError, TypeError):
            continue
        if row_date == target_date:
            name = str(r[1] or "").strip()
            chain_flag = str(r[2] or "").strip() if len(r) > 2 else ""
            is_chain = bool(chain_flag)  # 何か入力されてれば True
            if not name:
                return None
            return Spot(name=name, is_chain=is_chain, source="input")
    return None


def load_master_active(wb) -> list[Spot]:
    """マスタシートから status=active のスポットを全件取得

    シート構造（6列・2026-05-07・カテゴリ列追加版）:
        A: スポット名
        B: カテゴリ（社長参考用・記事生成では使わない）
        C: チェーン店?（空欄でない=チェーン店扱い）
        D: ステータス（active / paused / removed）
        E: 最終使用日
        F: 使用回数
    """
    rows = _get_rows(wb, MASTER_SHEET) or []
    # データ部は行6（index 5）以降
    spots = []
    for r in rows[5:]:
        if not r or r[0] is None:
            continue
        name = str(r[0] or "").strip()
        # B列=カテゴリ（参照のみ・記事生成では使わない）
        chain_flag = str(r[2] or "").strip() if len(r) > 2 else ""
        is_chain = bool(chain_flag)
        status = str(r[3] or "active").strip().lower() if len(r) > 3 else "active"
        if status != "active":
            continue
        if not name:
            continue
        spots.append(Spot(name=name, is_chain=is_chain, source="master"))
    return spots


def load_recent_used_names(wb, target_date: date, days: int = 7) -> set[str]:
    """配信ログから直近 N 日に使ったスポット名を取得（同週NGチェック用）"""
    rows = _get_rows(wb, LOG_SHEET)
    if rows is None:
        return set()
    # 配信ログのヘッダー: 行3 (index 2) - 配信日, 曜日, TOP1, TOP2, TOP3, 最下位, ラッキースポット
    # ラッキースポット列の値を target_date 直近days日分から拾う
    used = set()
    cutoff = target_date - timedelta(days=days)
    for r in rows[3:]:
        if not r or r[0] is None:
            continue
        try:
            row_date = to_date(r[0])
        except (ValueError, TypeError):
            continue
        if cutoff <= row_date < target_date:
            # 7列目（index 6）が「ラッキースポット」
            if len(r) > 6 and r[6]:
                used.add(str(r[6]).strip())
    return used


def select_lucky_spot(target_date, *, recent_used_extra: set[str] | None = None,
                     wb=None, seed: int | None = None) -> Spot:
    """target_date のラッキースポットを選定する

    Args:
        target_date: 配信対象日
        recent_used_extra: シミュレーション時に追加で除外したいスポット名集合
        wb: 既に開いている Workbook（シミュレーションで再利用するため）
        seed: ランダムシード（再現性が必要な場合）

    Returns:
        Spot
    """
    target_date = to_date(target_date)
    if seed is not None:
        random.seed(seed)
    # 2026-05-25 案A：xlsx フォールバック完全削除。Sheets 必須。
    # Sheets 取得失敗時は _get_rows() で None が返り、後段でエラー停止する。
    # ローカル CLI でテストする時のみ、明示的に wb を渡せば xlsx も使える。

    # Step 1: 入力シート検索（社長指名・最優先）
    spot = lookup_input_sheet(wb, target_date)
    if spot is not None:
        return spot

    # Step 2: マスタからフォールバック選定
    candidates = load_master_active(wb)

    # 直近7日に使ったスポットは除外
    used = load_recent_used_names(wb, target_date, days=7)
    if recent_used_extra:
        used = used | recent_used_extra
    candidates = [s for s in candidates if s.name not in used]

    if not candidates:
        # 候補ゼロ（極端なケース）→ 全件から選び直し
        candidates = load_master_active(wb)
        if not candidates:
            raise RuntimeError(f"マスタに active スポットが1件もありません ({target_date})")

    # シンプル化方針：tier 比率調整は廃止（2026-05-07 社長判断）
    # is_chain ベースで「直近7日でチェーン店ばかりなら個人店優先」程度の軽い調整のみ
    chain_ratio_recent = sum(1 for s in load_master_active(wb)
                             if s.name in used and s.is_chain) / max(len(used), 1)
    if chain_ratio_recent > 0.5:
        # 直近でチェーン店過剰 → 非チェーンを優先
        filtered = [s for s in candidates if not s.is_chain]
        if filtered:
            candidates = filtered
    elif chain_ratio_recent < 0.2 and len(used) >= 3:
        # 直近でチェーン店少なすぎ → チェーンも混ぜる
        pass  # そのまま全候補から選ぶ

    selected = random.choice(candidates)
    selected.source = "fallback"
    return selected


# ============================================================
# CLI / シミュレーション
# ============================================================

def simulate(start: date, end: date, seed: int = 42):
    """期間内の選定結果を一気に表示（dry-run）"""
    random.seed(seed)
    wb = load_workbook_ro()

    days = (end - start).days + 1
    results = []
    # シミュレーション中は配信ログがないので、直近選定スポットを集合で管理
    simulated_used: list[tuple[date, str]] = []

    for i in range(days):
        d = start + timedelta(days=i)
        # シミュレーション内で「直近7日に使った」を簡易計算
        recent_extra = {n for dt, n in simulated_used if d - dt <= timedelta(days=7)}
        spot = select_lucky_spot(d, wb=wb, recent_used_extra=recent_extra)
        results.append((d, spot))
        simulated_used.append((d, spot.name))

    # 出力
    print(f"\n=== 選定シミュレーション ({start} 〜 {end} / 計{days}日) ===\n")
    print(f"{'日付':12s}  {'曜':3s}  {'src':8s}  {'チェーン':6s}  {'name':30s}")
    print("-" * 80)
    for d, s in results:
        wd = WEEKDAY_JP[d.weekday()]
        src_label = {"input": "👤入力", "master": "🎲master", "fallback": "🎲rand"}.get(s.source, s.source)
        chain_mark = "● " if s.is_chain else "  "
        print(f"{d.isoformat():12s}  {wd:3s}  {src_label:8s}  {chain_mark:6s}  {s.name:30s}")

    # チェーン店比率
    n_chain = sum(1 for _, s in results if s.is_chain)
    n_named = len(results) - n_chain
    src_counter = Counter(s.source for _, s in results)
    print(f"\n--- 表記方式 ---")
    print(f"  指名表記         : {n_named:3d} 件 ({n_named*100//max(len(results),1)}%)")
    print(f"  「お近くの」表記 : {n_chain:3d} 件 ({n_chain*100//max(len(results),1)}%)")

    print(f"\n--- ソース集計 ---")
    for src, c in src_counter.most_common():
        label = {"input": "社長指名", "master": "マスタ選定", "fallback": "マスタフォールバック"}.get(src, src)
        print(f"  {label}: {c} 件")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", help="単日選定の対象日（YYYY-MM-DD）")
    parser.add_argument("--simulate", action="store_true", help="期間シミュレーション")
    parser.add_argument("--start", help="シミュレーション開始日（YYYY-MM-DD）")
    parser.add_argument("--end", help="シミュレーション終了日（YYYY-MM-DD）")
    parser.add_argument("--seed", type=int, default=42, help="ランダムシード")
    args = parser.parse_args()

    if args.simulate:
        if not args.start or not args.end:
            parser.error("--simulate には --start と --end が必要です")
        start = to_date(args.start)
        end = to_date(args.end)
        simulate(start, end, seed=args.seed)
        return

    if args.date:
        d = to_date(args.date)
        spot = select_lucky_spot(d, seed=args.seed)
        print(f"\n選定結果: {d} ({WEEKDAY_JP[d.weekday()]})")
        print(f"  name     : {spot.name}")
        print(f"  is_chain : {spot.is_chain}")
        print(f"  src      : {spot.source}")
        return

    parser.print_help()


if __name__ == "__main__":
    main()
